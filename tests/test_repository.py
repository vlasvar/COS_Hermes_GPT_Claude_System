import json
import io
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]


class RepositoryContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.schema = json.loads((ROOT / "schema/sheets.json").read_text(encoding="utf-8"))

    def test_core_tabs_are_unique(self):
        names = [tab["name"] for tab in self.schema["tabs"]]
        self.assertEqual(len(names), len(set(names)))

    def test_every_tab_has_a_stable_key(self):
        for tab in self.schema["tabs"]:
            with self.subTest(tab=tab["name"]):
                self.assertTrue(tab["key"].endswith("ID"))
                self.assertIn(tab["key"], tab["columns"])

    def test_every_agent_has_adapter_and_setup(self):
        expected = {
            "hermes": "HERMES.md",
            "chatgpt": "PROJECT_INSTRUCTIONS.md",
            "claude": "CLAUDE.md",
        }
        for adapter, instruction_file in expected.items():
            with self.subTest(adapter=adapter):
                folder = ROOT / "adapters" / adapter
                self.assertTrue((folder / "README.md").is_file())
                self.assertTrue((folder / instruction_file).is_file())

    def test_google_sheets_provisioner_matches_schema(self):
        script = (ROOT / "templates/google-sheets/Code.gs").read_text(encoding="utf-8")
        self.assertIn("@OnlyCurrentDoc", script)
        self.assertNotIn(".clear()", script)
        for tab in self.schema["tabs"]:
            with self.subTest(tab=tab["name"]):
                self.assertIn(f"name: '{tab['name']}'", script)
                for column in tab["columns"]:
                    self.assertIn(f"'{column}'", script)

    def test_finance_tabs_lead_the_schema(self):
        names = [tab["name"] for tab in self.schema["tabs"]]
        self.assertEqual(names[:4], ["System Check", "Expenses", "Budget", "Recurring Costs"])

    def test_starter_workbook_matches_schema(self):
        workbook_path = ROOT / "starter-kit" / "COS_DATABASE_TEMPLATE.xlsx"
        workbook = load_workbook(workbook_path, read_only=False, data_only=False)
        expected_names = ["START HERE", *[tab["name"] for tab in self.schema["tabs"]]]
        self.assertEqual(workbook.sheetnames, expected_names)

        for tab in self.schema["tabs"]:
            with self.subTest(tab=tab["name"]):
                sheet = workbook[tab["name"]]
                headers = [sheet.cell(1, index).value for index in range(1, len(tab["columns"]) + 1)]
                self.assertEqual(headers, tab["columns"])
                self.assertEqual(sheet.freeze_panes, "A2")

        system_check = workbook["System Check"]
        self.assertEqual(system_check["A2"].value, "CHECK-READ")
        self.assertEqual(system_check["A3"].value, "CHECK-WRITE")
        self.assertEqual(system_check["A4"].value, "CHECK-LOG")

    def test_distributable_zip_contains_complete_starter(self):
        archive_path = ROOT / "dist" / "COS_Finance_First_Starter.zip"
        required = {
            "COS_Finance_First_Starter/00_START_HERE.md",
            "COS_Finance_First_Starter/01_COPY_THIS_PROMPT.txt",
            "COS_Finance_First_Starter/COS_DATABASE_TEMPLATE.xlsx",
            "COS_Finance_First_Starter/Inbox/Expenses-and-Receipts/UPLOAD_HERE.txt",
            "COS_Finance_First_Starter/System/AGENT_RULES.md",
            "COS_Finance_First_Starter/System/FINANCE_WORKFLOW.md",
            "COS_Finance_First_Starter/System/DATA_DICTIONARY.md",
            "COS_Finance_First_Starter/System/OPTIONAL_PROFILE.md",
            "COS_Finance_First_Starter/System/Advanced/OPTIONAL_Code.gs",
        }
        with zipfile.ZipFile(archive_path) as archive:
            members = set(archive.namelist())
            self.assertTrue(required.issubset(members), required - members)
            workbook_bytes = archive.read(
                "COS_Finance_First_Starter/COS_DATABASE_TEMPLATE.xlsx"
            )
        workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True)
        self.assertIn("Expenses", workbook.sheetnames)
        self.assertIn("Budget", workbook.sheetnames)

    def test_bootstrap_prompt_requires_verified_write_capability(self):
        prompt = (ROOT / "starter-kit" / "01_COPY_THIS_PROMPT.txt").read_text(encoding="utf-8")
        self.assertIn("[PASTE GOOGLE DRIVE FOLDER LINK HERE]", prompt)
        self.assertIn("write", prompt.lower())
        self.assertIn("read it back", prompt.lower())
        self.assertIn("Workspace Operator", prompt)
        self.assertIn("identity details are optional", prompt.lower())

    def test_optional_apps_script_in_starter_matches_template(self):
        source = (ROOT / "templates" / "google-sheets" / "Code.gs").read_bytes()
        packaged_source = (
            ROOT / "starter-kit" / "System" / "Advanced" / "OPTIONAL_Code.gs"
        ).read_bytes()
        self.assertEqual(packaged_source, source)

    def test_default_configuration_uses_workspace_operator(self):
        config = (ROOT / "config" / "system.example.yaml").read_text(encoding="utf-8")
        self.assertIn('workspace_role: "operator"', config)
        self.assertIn("default_level: 2", config)
        self.assertIn("identity_required: false", config)

    def test_public_tree_has_no_private_instance_directory(self):
        self.assertFalse((ROOT / "private").exists())
        self.assertFalse((ROOT / "instance").exists())

    def test_instance_generator_creates_each_adapter_bundle(self):
        expected_instruction = {
            "hermes": "AGENTS.md",
            "chatgpt": "PROJECT_INSTRUCTIONS.md",
            "claude": "CLAUDE.md",
        }
        with tempfile.TemporaryDirectory() as temporary_root:
            for adapter, instruction_file in expected_instruction.items():
                with self.subTest(adapter=adapter):
                    target = Path(temporary_root) / adapter
                    result = subprocess.run(
                        [
                            sys.executable,
                            str(ROOT / "scripts/create_instance.py"),
                            str(target),
                            "--adapter",
                            adapter,
                        ],
                        capture_output=True,
                        text=True,
                        check=False,
                    )
                    self.assertEqual(result.returncode, 0, result.stderr)
                    self.assertTrue((target / instruction_file).is_file())
                    self.assertTrue((target / "USER_PROFILE.md").is_file())
                    self.assertTrue((target / "system.yaml").is_file())
                    for kernel_file in ("SYSTEM.md", "CONTEXT.md", "PERMISSIONS.md", "FINANCE_FIRST.md", "WORKFLOWS.md"):
                        self.assertTrue((target / kernel_file).is_file())

    def test_instance_generator_rejects_target_inside_repository(self):
        target = ROOT / "instance-generator-must-reject"
        result = subprocess.run(
            [
                sys.executable,
                str(ROOT / "scripts/create_instance.py"),
                str(target),
                "--adapter",
                "hermes",
            ],
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertNotEqual(result.returncode, 0)
        self.assertFalse(target.exists())


if __name__ == "__main__":
    unittest.main()
