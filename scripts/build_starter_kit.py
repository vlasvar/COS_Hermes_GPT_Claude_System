#!/usr/bin/env python
"""Build the finance-first workbook and distributable starter ZIP."""
from __future__ import annotations

import json
import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
STARTER = ROOT / "starter-kit"
DIST = ROOT / "dist"
WORKBOOK = STARTER / "COS_DATABASE_TEMPLATE.xlsx"
ARCHIVE = DIST / "COS_Finance_First_Starter.zip"
SCHEMA_PATH = ROOT / "schema" / "sheets.json"
OPTIONAL_SCRIPT = STARTER / "System" / "Advanced" / "OPTIONAL_Code.gs"

NAVY = "172033"
BLUE = "2855D9"
PALE_BLUE = "EAF0FF"
PALE_GREEN = "E7F6EE"
PALE_YELLOW = "FFF4CC"
WHITE = "FFFFFF"
DARK = "1F2937"
BORDER = "CFD7E6"

FINANCE_TABS = {"System Check", "Income", "Expenses", "Budget", "Recurring Costs"}
DATE_TOKENS = ("Date", "At", "Period Start", "Period End", "Timestamp")
AMOUNT_TOKENS = ("Amount", "Variance")


def set_header_style(sheet, columns: list[str]) -> None:
    for index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=name)
        cell.font = Font(color=WHITE, bold=True)
        cell.fill = PatternFill("solid", fgColor=BLUE if sheet.title in FINANCE_TABS else NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    sheet.sheet_view.showGridLines = False

    for index, name in enumerate(columns, start=1):
        width = max(13, min(34, len(name) + 5))
        if any(token in name for token in ("Description", "Summary", "Rationale", "Details", "Notes")):
            width = 30
        if "Link" in name:
            width = 28
        sheet.column_dimensions[get_column_letter(index)].width = width


def add_dropdown(sheet, column_name: str, values: list[str], max_row: int = 2000) -> None:
    headers = [cell.value for cell in sheet[1]]
    if column_name not in headers:
        return
    index = headers.index(column_name) + 1
    validation = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    validation.error = "Choose a value from the list."
    validation.errorTitle = "Invalid value"
    sheet.add_data_validation(validation)
    validation.add(f"{get_column_letter(index)}2:{get_column_letter(index)}{max_row}")


def format_columns(sheet, columns: list[str]) -> None:
    for index, name in enumerate(columns, start=1):
        letter = get_column_letter(index)
        if any(token in name for token in AMOUNT_TOKENS):
            for row in range(2, 2001):
                sheet[f"{letter}{row}"].number_format = '#,##0.00'
        elif name == "Confidence":
            add_dropdown(sheet, name, ["High", "Medium", "Low"])
        elif any(name == token or name.endswith(f" {token}") for token in DATE_TOKENS):
            for row in range(2, 2001):
                sheet[f"{letter}{row}"].number_format = "yyyy-mm-dd"


def add_start_sheet(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "START HERE"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:F1")
    title = sheet["A1"]
    title.value = "Chief of Staff — Finance-First Database"
    title.font = Font(size=20, bold=True, color=WHITE)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 42

    rows = [
        (3, "1", "Keep this workbook in the uploaded private Google Drive folder."),
        (4, "2", "Open it with Google Sheets and save it as a native Google Sheet named COS_DATABASE."),
        (5, "3", "Paste the folder link into the prompt in 01_COPY_THIS_PROMPT.txt."),
        (6, "4", "The agent must pass the System Check write-and-read-back test before onboarding."),
        (7, "5", "Describe income and expenses in plain English, including amounts and dates. Evidence can follow later."),
    ]
    for row, number, text in rows:
        sheet[f"A{row}"] = number
        sheet[f"A{row}"].font = Font(size=14, bold=True, color=BLUE)
        sheet[f"B{row}"] = text
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        sheet[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[row].height = 34

    sheet["A9"] = "Safety boundary"
    sheet["A9"].font = Font(bold=True, color=DARK)
    sheet.merge_cells("B9:F10")
    sheet["B9"] = (
        "The agent may maintain records inside this workspace. It must ask before spending money, "
        "sending messages, publishing, changing access, signing agreements, or deleting evidence."
    )
    sheet["B9"].fill = PatternFill("solid", fgColor=PALE_YELLOW)
    sheet["B9"].alignment = Alignment(wrap_text=True, vertical="center")

    sheet.column_dimensions["A"].width = 18
    for column in "BCDEF":
        sheet.column_dimensions[column].width = 18
    sheet.freeze_panes = "A3"


def build_workbook(schema: dict) -> None:
    workbook = Workbook()
    fixed_time = datetime(2026, 1, 1, tzinfo=timezone.utc)
    workbook.properties.title = "Chief of Staff Finance-First Database"
    workbook.properties.subject = "Generic user-owned finance-first Chief of Staff template"
    workbook.properties.creator = "Chief of Staff System contributors"
    workbook.properties.description = "Contains generic empty schemas and no personal data."
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    add_start_sheet(workbook)

    for tab in schema["tabs"]:
        sheet = workbook.create_sheet(tab["name"])
        columns = tab["columns"]
        set_header_style(sheet, columns)
        format_columns(sheet, columns)
        sheet.sheet_properties.tabColor = "2AA876" if tab["name"] in FINANCE_TABS else BLUE

        if tab["name"] == "System Check":
            checks = [
                ["CHECK-READ", "Read this row", "Pending", "", "Agent must confirm it can read the workbook."],
                ["CHECK-WRITE", "Write and read back", "Pending", "", "Replace this detail with a harmless test value, then read it back."],
                ["CHECK-LOG", "Write Agent Log", "Pending", "", "Add and verify an Agent Log entry."],
            ]
            for row_index, row in enumerate(checks, start=2):
                for column_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index, column=column_index, value=value)
            sheet.conditional_formatting.add(
                "C2:C2000",
                FormulaRule(formula=['C2="Passed"'], fill=PatternFill("solid", fgColor=PALE_GREEN)),
            )

        if tab["name"] == "Expenses":
            add_dropdown(sheet, "Status", ["Provisional", "Confirmed", "Rejected", "Duplicate"])
            add_dropdown(sheet, "Recurrence", ["One-off", "Recurring", "Possible", "Unknown"])
        elif tab["name"] == "Budget":
            add_dropdown(sheet, "Status", ["Draft", "Active", "Closed"])
        elif tab["name"] == "Recurring Costs":
            add_dropdown(sheet, "Status", ["Possible", "Active", "Paused", "Ended"])
            add_dropdown(sheet, "Frequency", ["Weekly", "Monthly", "Quarterly", "Annual", "Other", "Unknown"])

    WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(WORKBOOK)


def write_archive() -> list[str]:
    DIST.mkdir(parents=True, exist_ok=True)
    files = sorted(path for path in STARTER.rglob("*") if path.is_file())
    members: list[str] = []
    with zipfile.ZipFile(ARCHIVE, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        for path in files:
            relative = Path("COS_Finance_First_Starter") / path.relative_to(STARTER)
            info = zipfile.ZipInfo(relative.as_posix(), date_time=(2026, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o644 << 16
            archive.writestr(info, path.read_bytes())
            members.append(relative.as_posix())
    return members


def main() -> int:
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    build_workbook(schema)
    OPTIONAL_SCRIPT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(ROOT / "templates" / "google-sheets" / "Code.gs", OPTIONAL_SCRIPT)
    members = write_archive()
    print(json.dumps({
        "workbook": str(WORKBOOK.relative_to(ROOT)),
        "archive": str(ARCHIVE.relative_to(ROOT)),
        "sheets": ["START HERE", *[tab["name"] for tab in schema["tabs"]]],
        "archive_members": len(members),
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
